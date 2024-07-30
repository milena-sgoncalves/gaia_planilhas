from office365_api import SharePoint
import sys
import csv
from pathlib import PurePath
from openpyxl import Workbook

# 1 args = The name of the SharePoint Document Library
SHAREPOINT_FOLDER_NAME = sys.argv[1]
# 2 args = Export Type.  it can be "Excel" or "CSV"
EXPORT_TYPE = sys.argv[2]
# 3 args = Local Directory Path to save list dataset
DIR_PATH = sys.argv[3]
# 4 args = The name of the file that will get saved on local directory
FILE_NAME = sys.argv[4]

def set_file_ext(file_name, export_type):
    if export_type == 'Excel':
        file_name_with_ext = '.'.join([file_name, 'xlsx'])
    elif export_type == 'CSV':
        file_name_with_ext = '.'.join([file_name, 'csv'])
    else:
        file_name_with_ext = file_name
    return file_name_with_ext

def download_folder(folder_name, export_type, dir_path, file_name):
    sp = SharePoint()
    files_list = sp.get_file_properties_from_folder(folder_name)
    if export_type == 'Excel':
        save_to_excel(files_list, dir_path, file_name)
    elif export_type == 'CSV':
        save_to_csv(files_list, dir_path, file_name)
    else:
        print('Export type is not a valid type')

def save_to_csv(files_list, dir_path, file_name):
    dir_file_path = PurePath(dir_path, file_name)
    with open(dir_file_path, 'w', newline='\n', encoding='utf-8') as f:
        if files_list:
            header = files_list[0].keys()
            w = csv.DictWriter(f, header)
            w.writeheader()
            for file in files_list:
                w.writerow(file)

def save_to_excel(files_list, dir_path, file_name):
    dir_file_path = PurePath(dir_path, file_name)
    wb = Workbook()
    ws = wb.active
    # list of header names from SharePoint Document Library
    if files_list:
        header = files_list[0].keys()
        # write headers on first row
        for idx, name in enumerate(header):
            ws.cell(row=1, column=idx + 1, value=name)
        # write line items starting on second row
        row = 2
        for file in files_list:
            for idx, item in enumerate(file.items()):
                ws.cell(row=row, column=idx + 1, value=item[1])
            row += 1
    wb.save(dir_file_path)

if __name__ == '__main__':
    file_name = set_file_ext(FILE_NAME, EXPORT_TYPE)
    download_folder(SHAREPOINT_FOLDER_NAME, EXPORT_TYPE, DIR_PATH, file_name)
